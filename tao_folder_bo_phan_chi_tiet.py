import os
import re
import sys
import locale
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

# Cấu hình UTF-8 & Vietnamese locale
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except Exception:
        try:
            locale.setlocale(locale.LC_ALL, 'vi_VN')
        except Exception:
            pass

input_file = r'd:\DSNV\VTM-DSNV-07.2026.xlsx'
base_out_dir = r'd:\DSNV\BoPhan_ChiTiet'

def get_vietnamese_ho_sort_key(full_name):
    if pd.isna(full_name) or not str(full_name).strip():
        return ()
    parts = str(full_name).strip().split()
    if not parts:
        return ()
    return tuple(locale.strxfrm(p.lower()) for p in parts)

def sanitize_filename(name):
    clean_name = re.sub(r'[\\/*?:"<>|]', '_', str(name))
    return clean_name.strip()

def export_excel_styled(df_data, title_str, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách"

    font_header_title = Font(name="Times New Roman", size=11, bold=True)
    font_main_title = Font(name="Times New Roman", size=13, bold=True)
    font_col_header = Font(name="Times New Roman", size=10, bold=True)
    font_data = Font(name="Times New Roman", size=10)
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    ws['A1'] = "BỘ Y TẾ"
    ws['A1'].font = font_header_title
    ws['A2'] = "Bệnh viện Bạch Mai"
    ws['A2'].font = font_header_title
    
    ws['A5'] = title_str.upper()
    ws['A5'].font = font_main_title
    ws['A5'].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(df_data.columns))

    # Header dòng 7
    for col_idx, col_name in enumerate(df_data.columns, start=1):
        cell = ws.cell(row=7, column=col_idx, value=str(col_name))
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[7].height = 28

    # Data dòng 8+
    for row_idx, row_data in enumerate(df_data.values, start=8):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, start=1):
            if pd.isna(val):
                display_val = ""
            elif isinstance(val, pd.Timestamp):
                display_val = val.strftime('%d/%m/%Y')
            else:
                display_val = val

            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.font = font_data
            cell.border = thin_border
            
            if col_idx in [1, 2, 3, 5, 6, 7, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(filepath)

def tao_folder_chi_tiet_bo_phan(file_in, out_base):
    os.makedirs(out_base, exist_ok=True)
    
    df = pd.read_excel(file_in, header=7)
    df = df.dropna(subset=['Họ và tên']).copy()

    col_bo_phan = 'Bộ phận'
    col_chuc_danh = 'Tên chức danh nghề nghiệp'
    col_ho_ten = 'Họ và tên'

    bp_groups = df.groupby(col_bo_phan)
    
    total_folders = 0
    total_files = 0

    print(f"Bắt đầu khởi tạo folder cho {len(bp_groups)} bộ phận...\n")

    for bp_name, bp_df in bp_groups:
        total_folders += 1
        safe_bp_folder = sanitize_filename(bp_name)
        folder_path = os.path.join(out_base, safe_bp_folder)
        os.makedirs(folder_path, exist_ok=True)

        # 1. Tạo file tổng quát cho cả Bộ phận
        df_bp_sorted = bp_df.copy()
        df_bp_sorted['key_ho_ten'] = df_bp_sorted[col_ho_ten].apply(get_vietnamese_ho_sort_key)
        df_bp_sorted.sort_values(by='key_ho_ten', inplace=True)
        df_bp_sorted.drop(columns=['key_ho_ten'], inplace=True)
        df_bp_sorted['STT'] = range(1, len(df_bp_sorted) + 1)

        file_tong = os.path.join(folder_path, "00_DANH_SACH_TONG.xlsx")
        export_excel_styled(df_bp_sorted, f"DANH SÁCH TOÀN BỘ NHÂN VIÊN - {bp_name}", file_tong)
        total_files += 1

        # 2. Tách theo từng Chức danh trong Bộ phận này
        cd_groups = bp_df.groupby(col_chuc_danh, dropna=False)
        sub_files_count = 0

        for cd_name, cd_df in cd_groups:
            cd_str = "Chua_Phan_Loai" if pd.isna(cd_name) else str(cd_name)
            safe_cd_file = sanitize_filename(cd_str) + ".xlsx"
            file_cd_path = os.path.join(folder_path, safe_cd_file)

            df_cd_sorted = cd_df.copy()
            df_cd_sorted['key_ho_ten'] = df_cd_sorted[col_ho_ten].apply(get_vietnamese_ho_sort_key)
            df_cd_sorted.sort_values(by='key_ho_ten', inplace=True)
            df_cd_sorted.drop(columns=['key_ho_ten'], inplace=True)
            df_cd_sorted['STT'] = range(1, len(df_cd_sorted) + 1)

            export_excel_styled(df_cd_sorted, f"{bp_name} - CHỨC DANH: {cd_str}", file_cd_path)
            total_files += 1
            sub_files_count += 1

        print(f"📁 [{total_folders:2d}/13] Folder: '{safe_bp_folder}' -> Đã tạo {sub_files_count} file chức danh + 1 file tổng")

    print(f"\n-> TỔNG CỘNG ĐÃ TẠO {total_folders} FOLDER VÀ {total_files} FILE EXCEL TẠI: '{out_base}'")

if __name__ == '__main__':
    tao_folder_chi_tiet_bo_phan(input_file, base_out_dir)
