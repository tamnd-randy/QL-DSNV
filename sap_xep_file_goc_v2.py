import sys
import locale
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

# Cấu hình UTF-8 & Vietnamese locale
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except Exception:
        try:
            locale.setlocale(locale.LC_ALL, 'vi_VN')
        except Exception:
            pass

input_file = r'd:\DSNV\VTM-DSNV-07.2026.xlsx'
output_file = r'd:\DSNV\DSNV_SapXep_DonVi_HoTen.xlsx'
alt_output_file = r'd:\DSNV\VTM-DSNV-07.2026_SapXep.xlsx'

def get_vietnamese_ho_sort_key(full_name):
    if pd.isna(full_name) or not str(full_name).strip():
        return ()
    parts = str(full_name).strip().split()
    if not parts:
        return ()
    # Sắp xếp theo thứ tự HỌ -> ĐỆM -> TÊN (từ trái qua phải)
    return tuple(locale.strxfrm(p.lower()) for p in parts)

def sap_xep_va_xuat_excel(file_in, file_out):
    # 1. Đọc dữ liệu từ file gốc
    df = pd.read_excel(file_in, header=7)
    
    # Loại bỏ các dòng trống không có tên nhân viên
    df = df.dropna(subset=['Họ và tên']).copy()

    # Tìm tên các cột chính
    col_don_vi = 'Đơn vị' if 'Đơn vị' in df.columns else None
    col_bo_phan = 'Bộ phận' if 'Bộ phận' in df.columns else None
    col_ho_ten = 'Họ và tên' if 'Họ và tên' in df.columns else None

    print(f"Các cột sử dụng: Đơn vị='{col_don_vi}', Bộ phận='{col_bo_phan}', Họ và tên='{col_ho_ten}'")

    # 2. Tạo các khóa sắp xếp dạng tuple theo HỌ (từ trái qua phải)
    df['key_don_vi'] = df[col_don_vi].apply(lambda x: locale.strxfrm(str(x).strip().lower()) if pd.notna(x) else '')
    df['key_bo_phan'] = df[col_bo_phan].apply(lambda x: locale.strxfrm(str(x).strip().lower()) if pd.notna(x) else '')
    df['key_ho_ten'] = df[col_ho_ten].apply(get_vietnamese_ho_sort_key)

    # Sắp xếp theo ưu tiên: Đơn vị -> Bộ phận -> Họ và tên theo ABC của HỌ
    df.sort_values(by=['key_don_vi', 'key_bo_phan', 'key_ho_ten'], inplace=True)

    # Đánh lại STT liên tục từ 1 đến N
    if 'STT' in df.columns:
        df['STT'] = range(1, len(df) + 1)

    # Loại bỏ các cột phụ dùng để sort
    df.drop(columns=['key_don_vi', 'key_bo_phan', 'key_ho_ten'], inplace=True)

    # 3. Tạo Workbook mới với openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách nhân viên"

    # Định dạng Font & Align
    font_header_title = Font(name="Times New Roman", size=11, bold=True)
    font_main_title = Font(name="Times New Roman", size=16, bold=True)
    font_col_header = Font(name="Times New Roman", size=10, bold=True)
    font_data = Font(name="Times New Roman", size=10)
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Ghi khối tiêu đề đầu trang
    ws['A1'] = "BỘ Y TẾ"
    ws['A1'].font = font_header_title
    ws['A2'] = "Bệnh viện Bạch Mai"
    ws['A2'].font = font_header_title
    
    ws['A6'] = "DANH SÁCH NHÂN VIÊN (SẮP XẾP THEO ĐƠN VỊ - BỘ PHẬN - HỌ VÀ TÊN THEO ABC CỦA HỌ)"
    ws['A6'].font = font_main_title
    ws['A6'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(df.columns))

    # Ghi tên các cột tại dòng 8
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=8, column=col_idx, value=str(col_name))
        cell.font = font_col_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[8].height = 28

    # Ghi dữ liệu đã sắp xếp từ dòng 9
    for row_idx, row_data in enumerate(df.values, start=9):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, start=1):
            if pd.isna(val):
                display_val = ""
            elif isinstance(val, pd.Timestamp):
                display_val = val.strftime('%d/%m/%Y')
            else:
                display_val = val

            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.font = font_data
            cell.border = thin_border
            
            if col_idx in [1, 2, 3, 5, 6, 7, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Lưu file
    actual_saved = file_out
    try:
        wb.save(file_out)
        print(f"-> ĐÃ SẮP XẾP VÀ XUẤT THÀNH CÔNG: {len(df)} nhân viên.")
        print(f"-> File đầu ra: '{file_out}'")
    except PermissionError:
        print(f"[CẢNH BÁO] File '{file_out}' đang mở. Lưu thay thế vào '{alt_output_file}'")
        wb.save(alt_output_file)
        actual_saved = alt_output_file

    return actual_saved

if __name__ == '__main__':
    sap_xep_va_xuat_excel(input_file, output_file)
