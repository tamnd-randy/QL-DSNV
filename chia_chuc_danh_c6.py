import os
import re
import sys
import locale
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

# Cấu hình UTF-8 & Vietnamese locale
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except Exception:
        try:
            locale.setlocale(locale.LC_ALL, 'vi_VN')
        except Exception:
            pass

input_file = r'd:\DSNV\DSNV_Theo_BoPhan\Khoa Bệnh mạch máu - Tăng huyết áp và Tim mạch dự phòng (C6).xlsx'
output_dir = r'd:\DSNV\DSNV_Theo_BoPhan\Khoa_C6_Theo_ChucDanh'

def get_vietnamese_ho_sort_key(full_name):
    if pd.isna(full_name) or not str(full_name).strip():
        return ()
    parts = str(full_name).strip().split()
    if not parts:
        return ()
    return tuple(locale.strxfrm(p.lower()) for p in parts)

def sanitize_filename(filename):
    clean_name = re.sub(r'[\\/*?:"<>|]', '_', filename)
    return clean_name.strip()

def chia_c6_theo_chuc_danh(file_in, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    
    # Đọc dữ liệu từ file C6 (header tại dòng 7 trong openpyxl / 6 trong pandas 0-based)
    df = pd.read_excel(file_in, header=6)
    df = df.dropna(subset=['Họ và tên']).copy()

    col_chuc_danh = 'Tên chức danh nghề nghiệp'
    col_ho_ten = 'Họ và tên'

    # Nhóm theo Chức danh
    grouped = df.groupby(col_chuc_danh)

    font_header_title = Font(name="Times New Roman", size=11, bold=True)
    font_main_title = Font(name="Times New Roman", size=13, bold=True)
    font_col_header = Font(name="Times New Roman", size=10, bold=True)
    font_data = Font(name="Times New Roman", size=10)
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    created_files = []

    for chuc_danh_name, group_df in grouped:
        group_df = group_df.copy()
        
        # Sắp xếp theo Họ ABC
        group_df['key_ho_ten'] = group_df[col_ho_ten].apply(get_vietnamese_ho_sort_key)
        group_df.sort_values(by='key_ho_ten', inplace=True)
        group_df.drop(columns=['key_ho_ten'], inplace=True)
        
        # Đánh lại STT 1..N
        group_df['STT'] = range(1, len(group_df) + 1)

        safe_name = sanitize_filename(str(chuc_danh_name))
        filename = f"{safe_name}.xlsx"
        filepath = os.path.join(out_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách"

        ws['A1'] = "BỘ Y TẾ"
        ws['A1'].font = font_header_title
        ws['A2'] = "Bệnh viện Bạch Mai"
        ws['A2'].font = font_header_title
        
        ws['A5'] = f"KHOA C6 - DANH SÁCH {str(chuc_danh_name).upper()}"
        ws['A5'].font = font_main_title
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(group_df.columns))

        # Header dòng 7
        for col_idx, col_name in enumerate(group_df.columns, start=1):
            cell = ws.cell(row=7, column=col_idx, value=str(col_name))
            cell.font = font_col_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[7].height = 28

        # Data dòng 8+
        for row_idx, row_data in enumerate(group_df.values, start=8):
            ws.row_dimensions[row_idx].height = 20
            for col_idx, val in enumerate(row_data, start=1):
                if pd.isna(val):
                    display_val = ""
                elif isinstance(val, pd.Timestamp):
                    display_val = val.strftime('%d/%m/%Y')
                else:
                    display_val = val

                cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = font_data
                cell.border = thin_border
                
                if col_idx in [1, 2, 3, 5, 6, 7, 10, 11]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        wb.save(filepath)
        created_files.append((chuc_danh_name, len(group_df), filepath))

    print(f"\n-> TỔNG CỘNG ĐÃ TẠO {len(created_files)} FILE XEL THEO CHỨC DANH TẠI: '{out_dir}'")
    for cd, count, fp in created_files:
        print(f"  + [{count:2d} NV] {cd} -> {os.path.basename(fp)}")

if __name__ == '__main__':
    chia_c6_theo_chuc_danh(input_file, output_dir)
