import sys
import copy
import locale
import openpyxl
import pandas as pd

# Thiết lập Tiếng Việt chuẩn
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except Exception:
        try:
            locale.setlocale(locale.LC_ALL, 'vi_VN')
        except Exception:
            pass

input_file = r'd:\DSNV\VTM-DSNV-07.2026.xlsx'
output_file = r'd:\DSNV\DSNV_SapXep_DonVi_HoTen.xlsx'
alt_output_file = r'd:\DSNV\VTM-DSNV-07.2026_SapXep.xlsx'

def get_vietnamese_name_sort_key(full_name):
    if not full_name:
        return []
    parts = str(full_name).strip().split()
    if not parts:
        return []
    # Quy tắc ABC Tiếng Việt: Tên chính (từ cuối) -> Họ (từ đầu) -> Đệm (các từ ở giữa)
    ordered = [parts[-1]] + [parts[0]] + parts[1:-1]
    return [locale.strxfrm(p.lower()) for p in ordered]

def get_sort_key_for_row(row_cells, col_don_vi_idx, col_bo_phan_idx, col_ho_ten_idx):
    val_don_vi = row_cells[col_don_vi_idx - 1].value or ""
    val_bo_phan = row_cells[col_bo_phan_idx - 1].value or ""
    val_ho_ten = row_cells[col_ho_ten_idx - 1].value or ""

    key_don_vi = locale.strxfrm(str(val_don_vi).strip().lower())
    key_bo_phan = locale.strxfrm(str(val_bo_phan).strip().lower())
    key_ho_ten = get_vietnamese_name_sort_key(val_ho_ten)

    return (key_don_vi, key_bo_phan, key_ho_ten)

def sap_xep_excel_goc(file_in, file_out):
    wb = openpyxl.load_workbook(file_in)
    sheet = wb.active

    # Xác định các cột trên dòng 8 (header)
    col_stt_idx = 1
    col_ho_ten_idx = 4
    col_bo_phan_idx = 13
    col_don_vi_idx = 14

    for c in range(1, sheet.max_column + 1):
        h_val = str(sheet.cell(row=8, column=c).value or "").strip().lower()
        if h_val == 'stt':
            col_stt_idx = c
        elif 'họ và tên' in h_val or 'họ tên' in h_val:
            col_ho_ten_idx = c
        elif h_val == 'bộ phận':
            col_bo_phan_idx = c
        elif h_val == 'đơn vị':
            col_don_vi_idx = c

    print(f"Cột STT: {col_stt_idx}, Họ tên: {col_ho_ten_idx}, Bộ phận: {col_bo_phan_idx}, Đơn vị: {col_don_vi_idx}")

    # Đọc dữ liệu từ dòng 9 trở đi
    data_rows = []
    for r in range(9, sheet.max_row + 1):
        row_cells = [sheet.cell(row=r, column=c) for c in range(1, sheet.max_column + 1)]
        # Kiểm tra dòng có dữ liệu hay không
        if any(cell.value is not None for cell in row_cells):
            data_rows.append(row_cells)

    print(f"Tổng số dòng dữ liệu đọc được: {len(data_rows)}")

    # Sắp xếp danh sách dòng dữ liệu
    data_rows.sort(key=lambda cells: get_sort_key_for_row(cells, col_don_vi_idx, col_bo_phan_idx, col_ho_ten_idx))

    # Đọc lại tất cả giá trị và định dạng để ghi đè vào sheet
    row_values_list = []
    for row_cells in data_rows:
        vals = [cell.value for cell in row_cells]
        row_values_list.append(vals)

    # Ghi lại các giá trị đã sắp xếp vào sheet từ dòng 9
    for r_idx, vals in enumerate(row_values_list, start=9):
        # Đánh lại STT liên tục
        vals[col_stt_idx - 1] = r_idx - 8
        
        for c_idx, val in enumerate(vals, start=1):
            sheet.cell(row=r_idx, column=c_idx).value = val

    # Lưu file kết quả
    try:
        wb.save(file_out)
        print(f"-> Đã sắp xếp và lưu thành công file mới tại: '{file_out}'")
    except PermissionError:
        print(f"[CẢNH BÁO] File '{file_out}' đang mở. Lưu thay thế tại '{alt_output_file}'")
        wb.save(alt_output_file)
        file_out = alt_output_file

    return file_out

if __name__ == '__main__':
    sap_xep_excel_goc(input_file, output_file)
