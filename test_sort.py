import sys
import openpyxl
import locale

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except Exception:
        pass

wb = openpyxl.load_workbook(r'd:\DSNV\VTM-DSNV-07.2026.xlsx')
sheet = wb.active

print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
for r in range(1, 10):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
